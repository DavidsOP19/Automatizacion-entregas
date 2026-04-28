import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

# ── 1. Cargar datos ──────────────────────────────────────────────
df = pd.read_csv("data/deliveries.csv", parse_dates=["date"])
print(f"✓ Datos cargados: {len(df)} registros")

# ── 2. Análisis con pandas ───────────────────────────────────────
total_orders = len(df)
delivered = len(df[df["status"] == "Entregado"])
delayed = len(df[df["status"] == "Retrasado"])
failed = len(df[df["status"] == "Fallido"])
delivery_rate = round((delivered / total_orders) * 100, 1)
avg_days = round(df[df["status"] == "Entregado"]["delivery_time_days"].mean(), 1)

region_summary = df.groupby("region").agg(
    total=("order_id", "count"),
    entregados=("status", lambda x: (x == "Entregado").sum()),
    retrasados=("status", lambda x: (x == "Retrasado").sum()),
    tiempo_promedio=("delivery_time_days", "mean")
).round(1).reset_index()

driver_summary = df.groupby("driver").agg(
    total=("order_id", "count"),
    entregados=("status", lambda x: (x == "Entregado").sum()),
    tasa_exito=("status", lambda x: round((x == "Entregado").sum() / len(x) * 100, 1))
).reset_index().sort_values("tasa_exito", ascending=False)

print(f"✓ Tasa de entrega: {delivery_rate}% | Tiempo promedio: {avg_days} días")

# ── 3. Exportar Excel base ───────────────────────────────────────
os.makedirs("output", exist_ok=True)
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
output_path = f"output/reporte_entregas_{timestamp}.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    df.to_excel(writer, sheet_name="Datos", index=False)
    region_summary.to_excel(writer, sheet_name="Por Region", index=False)
    driver_summary.to_excel(writer, sheet_name="Por Repartidor", index=False)

# ── 4. Formato profesional ───────────────────────────────────────
wb = load_workbook(output_path)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
BORDER_SIDE = Side(style="thin", color="BFBFBF")
BORDER      = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                     top=BORDER_SIDE, bottom=BORDER_SIDE)

def format_sheet(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = BORDER
            if i % 2 == 0:
                cell.fill = ALT_FILL
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 4

for sheet_name in ["Datos", "Por Region", "Por Repartidor"]:
    format_sheet(wb[sheet_name])

# ── 5. Hoja Resumen Ejecutivo ────────────────────────────────────
ws_res = wb.create_sheet("Resumen Ejecutivo", 0)
ws_res.sheet_view.showGridLines = False

metricas = [
    ("REPORTE DE ENTREGAS LOGÍSTICAS", ""),
    (f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", ""),
    ("", ""),
    ("Métrica", "Valor"),
    ("Total de pedidos", total_orders),
    ("Pedidos entregados", delivered),
    ("Pedidos retrasados", delayed),
    ("Pedidos fallidos", failed),
    ("Tasa de entrega (%)", delivery_rate),
    ("Tiempo promedio de entrega (días)", avg_days),
]

for row_idx, (label, value) in enumerate(metricas, start=1):
    ws_res.cell(row=row_idx, column=1, value=label)
    ws_res.cell(row=row_idx, column=2, value=value)

ws_res["A1"].font = Font(bold=True, size=14, color="1F4E79")
for col in ["A4", "B4"]:
    ws_res[col].fill = HEADER_FILL
    ws_res[col].font = HEADER_FONT
ws_res.column_dimensions["A"].width = 38
ws_res.column_dimensions["B"].width = 16

# ── 6. Gráfica de barras ─────────────────────────────────────────
ws_chart = wb["Por Region"]
chart = BarChart()
chart.type = "col"
chart.title = "Pedidos por Región"
chart.y_axis.title = "Cantidad"
chart.x_axis.title = "Región"
chart.style = 10
chart.width = 18
chart.height = 12

data_ref = Reference(ws_chart, min_col=2, max_col=3,
                     min_row=1, max_row=ws_chart.max_row)
cats_ref = Reference(ws_chart, min_col=1,
                     min_row=2, max_row=ws_chart.max_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws_chart.add_chart(chart, "F2")

wb.save(output_path)
print(f"✓ Reporte generado: {output_path}")
print("✓ Proceso completado exitosamente")
